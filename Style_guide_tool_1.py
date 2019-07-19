# -*- coding: utf-8 -*-
"""
Created on Thu Oct  4 14:51:21 2018

@author: a-whalen
"""

import openpyxl
import re
import tkinter as tk
import os
from openpyxl.styles import Alignment, PatternFill, Border, Side
from tkinter import messagebox, Button, Entry, END, Radiobutton
from tkinter.filedialog import askopenfilename

class App(tk.Frame):    
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        self.master = master
        self.initializeUI()
        
    def initializeUI(self):
        self.master.geometry('550x150')
        self.master.title('XXX様スタイルガイドツール')
        self.fileEntry = Entry(
                self.master, width=50)
        self.fileEntry.grid(row=0, column=1, padx=10, pady=10)
        self.getFile = Button(
                self.master, text='ファイルを選択', command=self.selectFile
                )
        self.getFile.grid(row=0, column=2)
        self.lang = tk.IntVar()
        self.lang.set(1)
        self.langEng = Radiobutton(self.master, text='英語', 
                                   variable=self.lang, value=1)
        self.langEng.grid(row=3, column=0)
        self.langChi = Radiobutton(self.master, text='簡体語',
                                   variable=self.lang, value=2)
        self.langChi.grid(row=3, column=1)
        self.exitBtn = Button(
                self.master,text='終了', command=self.quitApp
                )
        self.exitBtn.grid(row=4, column=2, pady=20)
        self.startApp = Button(
                self.master, text='実行', command=self.startCheck
                )
        self.startApp.grid(row=4, column=3, pady=20)
    
    def quitApp(self):
        global root
        root.destroy()
        
    def selectFile(self):
        filename = askopenfilename()
        self.fileEntry.delete(0,END)
        self.fileEntry.insert(0,filename)
        return
    
    def startCheck(self):
        wb = None
        filename = self.fileEntry.get()
        try:
            wb = openpyxl.load_workbook(filename)
            messagebox.showinfo('XXX様スタイルガイドツール', os.path.basename(filename) + 'を開きます。')
        except OSError:
            messagebox.showinfo('XXX様スタイルガイドツール', '正しいファイルパースを選択してください。')
            return 
       
        sheet = None
        
        try:
            sheet = wb['checkSheet']
        except KeyError:
            messagebox.showinfo('XXX様スタイルガイドツール', '英数チェックファイルを使ってください。')
            return
        
        maxRow = sheet.max_row
        titleName = 'Title.xlsx.sdlxliff'
        tableName = 'table'
        freeName = 'Free.xlsx.sdlxliff'
        imageName = '.xml'
        
        header = sheet.cell(row=1, column=15)
        header.value = 'XXX様'
        header.fill = PatternFill('solid', fgColor='BFBFBF')
        thin_border = Border(
           top=Side(border_style='thin', color='00000000'),
           right=Side(border_style='thin', color='00000000'),
           left=Side(border_style='thin', color='00000000'),
           bottom=Side(border_style='thin', color='00000000')
        )
        header.border = thin_border
        sheet.column_dimensions['O'].width = 30
        
        
        for i in range(2, maxRow+1):
            sheet.cell(row=i, column=15).value = None
            sheet.cell(row=i, column=15).alignment = Alignment(wrapText=True)    
            sheet.cell(row=i, column=15).border = thin_border
        
        def errors(errVal, word, case):
            if case == 'upper':
                errTxt = ' は大文字ではない'
            elif case == 'lower':
                errTxt = ' は小文字ではない'
            elif case == 'dot':
                errTxt = ' のユニコードが違う'
            elif case == 'singular':
                errTxt = ' は複数形ではない'
            elif case == 'mismatch':
                errTxt = ' の訳文が合っていない'
            elif case == 'delete':
                errTxt = ' は翻訳対象以外'
            if not errVal:
                errVal = word + errTxt
            else:
                errVal = err.value + '\n' + word + errTxt
            return errVal
        
        def isWord(word):
            nonCapWords = ["with", "per", "for", "the", "but", "from"]
            if not word.isalpha():
                return False
            elif len(word)<3:
                return False
            elif word in nonCapWords:
                return False
            else:
                return True
        
        def checkCount(string, src, tar):
            if src.count(string) == tar.count(string):
                return True
            else:
                return False
            
        def tokenize(string):
            punc = '!@#$%^&*()_-+={}[]:;"\'|<>,.?/~`'
            for marker in punc:
                text = string.replace(marker, " ")
            split = text.split()
            return split
       
        for i in range(2, maxRow+1):
            t = sheet.cell(row=i, column=1)
            source = sheet.cell(row=i, column=4)
            target = sheet.cell(row=i, column=6)
            err = sheet.cell(row=i, column=15)
            src = str(source.value)
            tar = str(target.value)
            if src and tar:
                #check 二硫化モリブデン
                if '二硫化モリブデン' in src and self.lang.get() ==1:
                    if src.count('二硫化モリブデン') != tar.count(r'MoS<sub>2</sub>'):
                        err.value = errors(err.value, 'MoS<sub>2</sub>', 'mismatch')
                #check dot unicode
                if self.lang.get() == 1:
                    if '·' in str(tar):
                        case = 'dot'
                        err.value = errors(err.value, '·', case)
                elif self.lang.get() == 2:
                    if '･' in tar:
                        case = 'dot'
                        err.value = errors(err.value, '･', case)
                #check arrows
                if re.compile('←\d').search(src):
                    if re.compile('←\d').search(src).group(0) not in tar:
                        err.value = errors(err.value, 
                                           re.compile('←\d').search(src).group(0),
                                           'mismatch')
                #check 全角 space
                if '　' in src:
                    if self.lang.get() ==1 and '  ' not in tar:
                        pos = src.find('　')
                        if src[pos] != 0:
                            starNum = re.compile(r"＊\d：　")
                            if not starNum.search(src):
                                err.value = errors(err.value, '全角スペース', 'mismatch')
                    if self.lang.get() ==2 and '　' not in tar:
                        err.value = errors(err.value, '全角スペース', 'mismatch')
                #check brackets
                if '[[' in src:
                    if src.count('[[') != tar.count('[['):
                        err.value = errors(err.value, '[[', 'mismatch')
                if ']]' in src:
                    if not checkCount(']]', src, tar):
                        err.value = errors(err.value, ']]', 'mismatch')
                #check tab character
                if r'\t' in src:
                    if not checkCount(r'\t', src, tar):
                        err.value = errors(err.value, r'\t', 'mismatch')
                #check hyphens
                if '----' in src:
                    if not checkCount('----', src, tar):
                        err.value = errors(err.value, '----', 'mismatch')
                #check slashes
                if '//' in src:
                    if not checkCount('//', src, tar):
                        err.value = errors(err.value, '//', 'mismatch')
                #check bullets
                if '●' in src:
                    if not checkCount('●', src, tar):
                        err.value = errors(err.value, '●', 'mismatch')
                #check maru suuji
                maruSuuji = '①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳'
                for num in maruSuuji:
                    if num in src:
                        if not checkCount(num, src, tar):
                            err.value = errors(err.value, num, 'mismatch')
                #check ＊1：, etc.
                if re.compile('＊\d：').search(src):
                    if re.compile('＊\d：').search(src).group(0) not in tar:
                        err.value = errors(err.value,
                                           re.compile('＊\d：').match(src).group(0),
                                           'mismatch')
                #check zero backslash
                if 'バックラッシュ０' in src or 'バックラッシュゼロ' in src or r'バックラッシュ<rubi>0@ゼロ</rubi>' in src:
                    if self.lang.get() == 1:
                        if 'zero backslash' not in tar:
                            err.value = errors(err.value, 'zero backslash', 'mismatch')
                    elif self.lang.get() == 2:
                        if '零背隙' not in tar:
                            err.value = errors(err.value, '零背隙', 'mismatch')
                #check dimensional table
                if '寸法・価格表' in src or '価格表' in src:
                    if self.lang.get() == 1:
                        if 'dimensional table' not in tar:
                            err.value = errors(err.value, 'dimensional table', 'mismatch')
                    elif self.lang.get() == 2:
                        if '尺寸表' not in tar:
                            err.value = errors(err.value, '尺寸表', 'mismatch')
                    #check 価格
                elif '価格' in src:
                    if '【データ上で消去】' not in tar:
                        err.value = '価格は翻訳対象以外'
                    else:
                        err.value = ''
                #check 単価:
                if '単価' in src:
                    if '【データ上で消去】' not in tar:
                        err.value = '単価は翻訳対象以外'
                    else:
                        err.value = ''                   
                #check 円
                if '円' in src:
                    if '【データ上で消去】' not in tar:
                        err.value = '円は翻訳対象以外'
                    else:
                        err.value= ''
                #check 納期
                if '納期' in src:
                    if '【データ上で消去】' not in tar:
                        err.value = '納期は翻訳対象以外'
                #check 出荷
                if '出荷' in src and '【データ上で消去】' not in tar:
                    err.value = errors(err.value, '出荷', 'delete')
            if self.lang.get() == 1 and t.value:
                #check free and table caps
                if freeName in t.value or tableName in t.value:
                    text = tokenize(tar)
                    case = 'upper'
                    for word in text:
                        if isWord(word) and word.islower():
                            err.value = errors(err.value, word, case) 
                #check title caps
                if titleName in t.value:
                    text = tokenize(tar)
                    case = 'upper'
                    for word in text:
                        if isWord(word) and word.islower():
                            err.value = errors(err.value, word, case)  
                #check image caps
                if imageName in t.value:
                    text = tokenize(tar)
                    case = 'lower'
                    if text[0].islower():
                        err.value = errors(err.value, text[0], case)
                    for word in text[1:]:
                        if isWord(word) and not word.islower():
                            err.value = errors(err.value, word, case)
        newFilename = 'Checked_' + os.path.basename(filename)
        wb.save(os.path.join(os.path.dirname(filename), newFilename))
        messagebox.showinfo('XXX様スタイルガイドツール', '終わりました！')   
            
if __name__ == '__main__':   
    root = tk.Tk()
    app = App(root)
    root.mainloop()              
