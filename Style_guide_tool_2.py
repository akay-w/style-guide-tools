# -*- coding: utf-8 -*-
"""
Created on Fri Nov  9 12:36:19 2018

@author: a-whalen
"""
import tkinter as tk
from tkinter import messagebox
from tkinter.filedialog import askopenfilename
import openpyxl
from openpyxl.styles import Alignment
import re
import os

#Create Tkinter GUI window
class App(tk.Frame):    
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        self.master = master
        self.initializeUI()
        
    def initializeUI(self):
        #GUI window dimensions
        self.master.geometry('500x100')
        #GUI window title
        self.master.title('XXX Style Guide Tool')
        #Blank for file name
        self.fileEntry = tk.Entry(
                self.master, width=50)
        self.fileEntry.grid(row=0, column=1, padx=10, pady=10)
        #Select file button
        self.getFile = tk.Button(
                self.master, text='Select File', command=self.selectFile
                )
        self.getFile.grid(row=0, column=2)
        #Exit app button
        self.exitBtn = tk.Button(
                self.master,text='Exit', command=self.quitApp,
                )
        self.exitBtn.grid(row=3, column=2, pady=20)
        #Run check button
        self.startApp = tk.Button(
                self.master, text='Run', command=self.startCheck,
                )        
        self.startApp.grid(row=3, column=3, pady=20)
    #Function called by exit button to quit app    
    def quitApp(self):
        global root
        root.destroy()
    #Function called by select file button to choose file    
    def selectFile(self):
        filename = askopenfilename()
        self.fileEntry.delete(0,tk.END)
        self.fileEntry.insert(0,filename)
        return
    #Function called by run button to run checks
    def startCheck(self):
        #Function for finding all indices of a given string in another string
        def find_all(a_str, sub):
            start = 0
            while True:
                start = a_str.find(sub, start)
                if start == -1: return
                yield start
                start += len(sub)     
        
        #Function for checking if a given segment is an individual word
        def check_if_single_word(tar, text):
            text_pos = list(find_all(tar, text))
            for i in range(0, len(text_pos)):
                            isSingleWord = True
                            index = text_pos[i]
                            if index > 0:
                                if tar[index-1].isalpha():
                                    isSingleWord = False
                            if (index + len(text)) <= len(tar)-1:
                                if tar[index + len(text)].isalpha():
                                    isSingleWord = False
            return isSingleWord
        
        def checkStyleRules(tar):
            errStr = ""
            styErr = "\nStyle error: "
            #lower case target for checks that ignore case
            lowTar = tar.casefold()
            
            #variable for any automatic replacements made
            subTar = tar
                    
            #if bullseye is in tar, replace     
            if u"\u25CE" in tar:
                errStr += styErr + u"\u25CE" + " changed to Very good"
                subTar = subTar.replace(u"\u25CE", "Very good")
            
            #if ○ or ○ in tar, produce error
            if "○" in tar or "○" in tar:
                #TODO: think about exceptions for replacement
                errStr += styErr + "○ incorrectly translated (chart)"
                
            #if triangle is in tar, replace       
            if "△" in tar:
                errStr += styErr + "△ incorrectly translated (chart)"   
                subTar = subTar.replace("△", "Limited use")
                
            #TODO: consider other possible contexts for replacement
            #also, should we check for spaces before/after for dimensions?
            #if × is the only character on the line, assume it's a chart        
            if "×" in tar:
                if len(tar) == 1:
                    errStr += styErr + "× incorrectly translated (chart)"
                    
            #if ! in target, replace  
            if "!" in tar:
                errStr += styErr + '! in target'
                subTar = subTar.replace("!", ".")
            
            #TODO: consider how to appropriately tackle / spacing
            #if "/ " in tar or " /" in tar:
                #errStr += styErr + 'Space before/after "/"'
            
            #Check for forbidden trademark/patent related words   
            #TODO: divide into replaceable and notreplaceable
            trademarks = ["™", "®", "trademark", "trademarked", "design registered", "PAT.P", "patent", "patented"]
            for item in trademarks:
                if item.casefold() in lowTar:
                    errStr += styErr + " forbidden word " + item + " in target"
            
            #if ～ in target, produce error
            if "～" in tar:
                errStr += styErr + '～ in target'
            
            #if ~ in target, produce error
            #Note: This isn't explicitly on the list but maybe leave it in?
            if "~" in tar:
                errStr += styErr + '~ in target' 
                
                                
            #Check for violations of Num V AC/DC format
            #eliminate all incorrect options:
            #DC12V, DC 12V, DC 12 V, 12VDC, 12 VDC (same idea for AC)
            #VDC and VAC are always wrong, so eliminate those first
            #TODO: automatically replace with correct format
            if "VDC" in tar or "VAC" in tar:
                errStr += styErr + "V DC or V AC format incorrect"
            #include this first condition to eliminate wasting time with regex
            if "DC" in tar or "AC" in tar:
                #check for DC12V, DC 12V, DC 12 V
                volts = re.compile(r"(AC|DC)(\s?)(\d+)(\s?)(V)")
                if volts.search(tar):
                    errStr += styErr + "V DC or V AC incorrect in target"
            
            #cultural sensitivities
            #TODO: come up with more forbidden words
            culturalNG = ["woman", "women", "elderly"]
            for item in culturalNG:
                if item.casefold() in lowTar:
                    errStr += styErr + " forbidden word " + item + " in target"
            
            #check for spaces before colons and replace
            #TODO: determine what other colon rules should be checked
            if " :" in tar:
                errStr += styErr + 'Space before ":"'
                subTar = subTar.replace(" :", ":")
                
            #Check for commas in 4 digit numbers  
            #TODO: determine if anything necessary for decimal separators?, replacement
            ketaKugiri = re.compile(r"\d{4}\d*")
            noComma = ketaKugiri.findall(tar)
            if len(noComma) > 0:
                for j in range(0, len(noComma)):
                    index = tar.find(noComma[j])
                    isSingleNum = True
                    if index > 0:
                        #Exclude numbers beginning/ending with <,-,> (tags)
                        if tar[index-1] == "-" or tar[index-1] == "<":
                            isSingleNum = False
                    if index + len(noComma[j]) <= len(tar)-1:
                        if tar[index + len(noComma[j])] == "-" or tar[index + len(noComma[j])] == ">":
                            isSingleNum = False
                    if isSingleNum == True:
                        errStr += styErr + "Missing comma in " + noComma[j]    
            
            #Check color list
            #TODO: expand list with commonly-used colors
            colors = ["red", "blue", "yellow", "green", "orange", "white", "black"]
            for color in colors:
                colorString = color + " color"
                if colorString.casefold() in lowTar:
                    errStr += styErr + 'Unneeded word "color" in ' + colorString
                    subTar = subTar.replace(colorString, color)
            
            #check to see if number of ( and ) are the same
            if tar.count("(") != tar.count(")"):
                errStr += styErr + "Number of opening/closing () does not match"
            
            #check for mismatched () even if 2byte
            if tar.count("（") != tar.count("）"):
                errStr += styErr + "Number of opening/closing () does not match"
                
            #check for spaces before % and replace
            if " %" in tar:
                errStr += styErr + 'Space before "%"'  
                subTar = subTar.replace(" %", "%")
                
            #check for space before ％ and replace (to still catch spacing even if 2byte char)                
            if " ％" in tar:
                errStr += styErr + 'Space before "％"'
                subTar = subTar.replace(" ％", "%")
                
            #check for space before ° and replace
            if " °" in tar:
                errStr += styErr + 'Space before "°"' 
                subTar = subTar.replace(" °", "°")

            #check for space before °C and replace
            if " °C" in tar:
                errStr += styErr + 'Space before "°C"'
                subTar = subTar.replace(" °C", "°C")
                
            #check for space even if 2byte and replace
            if " ℃" in tar:
                errStr += styErr + 'Space before "°C"'
                subTar = subTar.replace(" ℃", "°C")
            
            #check for space before phi and replace
            if " ø" in tar:
                errStr += styErr + 'Space before "ø"'
                subTar = subTar.replace(" ø", "ø")
            
            #check for space before phi even if 2byte and replace
            if " Ø" in tar:
                errStr += styErr + 'Space before "ø"'
                subTar = subTar.replace(" Ø", "ø")
                
            #Check for spaces between numbers and the following units and replace
            #unit list taken from pg12 of XXX style guide (feel free to add more)
            units = re.compile(r"(\d+)(cm|Hz|kHz|MHz|GHz|km|L|mL|MB|GB|g|kg|m|mm|in|lb)")
            unitMatch = units.findall(tar)
            if len(unitMatch) > 0:
                for i in range(0, len(unitMatch)):
                    num = unitMatch[i][0]
                    unit = unitMatch[i][1]
                    errStr += styErr + "No space between number and units (" + \
                    num + unit + ")"
                    subTar = subTar.replace(num + unit, num + " " + unit)
            
            #check for Velcro
            #TODO: replacement
            if "Velcro" in tar:
                errStr += styErr + "forbidden word Velcro"

            return errStr, subTar
        
        def checkGrammarRules(tar):
            graErr = "\nGrammar error: "
            errStr = ""
            #lower case target for checks that ignore case
            lowTar = tar.casefold()
            
            #check for contractions
            #TODO: add more cases, replacement
            if "n't" in tar:
                tokens = tar.split()
                for token in tokens:
                    if "n't" in token:
                        errStr += graErr + "contraction used (" + token + ")"
                        
            #flag instances of its
            if "its" in tar:
                isSingleWord = check_if_single_word(tar, "its")
                if isSingleWord:
                    errStr += graErr + "check its vs. it's"
            
            #flag instances of it's
            if "it's" in tar:
                isSingleWord = check_if_single_word(tar, "it's")
                if isSingleWord:
                    errStr += graErr + "check its vs. it's"
                    
            #check uncountable nouns
            #TODO: expand list
            uncountable_nouns = ["papers", "datas"]
            for noun in uncountable_nouns:
                if noun.casefold() in lowTar:
                    errStr += graErr + "change uncountable noun to singular (" + \
                    noun + ")"
            return errStr
        
        #get filename from GUI window blank
        filename = self.fileEntry.get()
        wb = None
        ws = None
        
        #open file in openpyxl
        try:
            #for xlsm functionality
            if filename.endswith(".xlsm"):
                wb = openpyxl.load_workbook(filename=filename, read_only=False, keep_vba=True)
            else:
                wb = openpyxl.load_workbook(filename)
            ws = wb.active
        #if file cannot be opened correctly, catch the error
        except:
            messagebox.showwarning("XXX Style Guide Tool",
                                   "Please select a valid file.")
            return
        
        #import data from sample sheet
        errStr = ""
        if wb and ws:
            maxRow = ws.max_row
            for i in range(2, maxRow+1):
                tar = str(ws.cell(row = i, column = 1).value)
                ws.cell(row = i, column = 3).value = ""
                errStr = ""
                #run checks and print subbed text to col 2 and errors to col 3
                if tar:
                    styleResults = checkStyleRules(tar)
                    errStr = styleResults[0]
                    subTar = styleResults[1]
                    grammarResults = checkGrammarRules(tar)
                    errStr += grammarResults
                    ws.cell(row = i, column = 2).value = subTar
                    ws.cell(row = i, column = 2).alignment = Alignment(wrapText=True)
                    ws.cell(row = i, column = 3).value = errStr
                    ws.cell(row = i, column = 3).alignment = Alignment(wrapText=True)
            
            #save checked file with new file name
            newFilename = 'Checked_' + os.path.basename(filename)
            #select same folder as original file and only change filename
            try:
                wb.save(os.path.join(os.path.dirname(filename), newFilename))
                messagebox.showinfo('XXX Style Guide Tool', 'Check finished!') 
            except PermissionError:
                messagebox.showwarning("XXX Style Guide Tool", "Please close all Excel files and try again.")

#initialize GUI for main thread
if __name__ == '__main__':   
    root = tk.Tk()
    app = App(root)
    root.mainloop()
