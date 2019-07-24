# -*- coding: utf-8 -*-

import openpyxl
import tkinter as tk
from tkinter import messagebox
from tkinter.filedialog import askopenfilename
import os

class App(tk.Frame):    
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        self.master = master
        self.initializeUI()
        
    def initializeUI(self):
        self.master.geometry('550x175')
        self.master.title('XXX様映像スタイルガイドツール')
        self.file_entry = tk.Entry(
                self.master, width=50)
        self.file_entry.grid(row=0, column=1, padx=10, pady=10)
        self.get_file = tk.Button(
                self.master, text='ファイルを選択', command=self.select_file
                )
        self.get_file.grid(row=0, column=2)
        self.start_cell = tk.Entry(
                self.master, width=50)
        self.start_cell.grid(row=1, column=1, padx=10, pady=10)
        self.start_cell_label = tk.Label(
                self.master, text='最初のセルを入力　例：G4')
        self.start_cell_label.grid(row=1, column=2)
        self.lang = tk.IntVar()
        self.lang.set(1)
        self.lang_eng = tk.Radiobutton(self.master, text='英語', 
                                   variable=self.lang, value=1)
        self.lang_eng.grid(row=3, column=1)
        self.lang_chi = tk.Radiobutton(self.master, text='簡体語',
                                   variable=self.lang, value=2)
        self.lang_chi.grid(row=3, column=2)
        self.exit_btn = tk.Button(
                self.master,text='終了', command=self.quit_app
                )
        self.exit_btn.grid(row=4, column=2, pady=20)
        self.start_app = tk.Button(
                self.master, text='実行', command=self.start_check
                )
        self.start_app.grid(row=4, column=3, pady=20)
    
    def quit_app(self):
        """
        Close GUI window
        """
        global root
        root.destroy()
        
    def select_file(self):
        """
        Prompts user to select file to check
        """
        filename = askopenfilename()
        self.file_entry.delete(0,tk.END)
        self.file_entry.insert(0,filename)
        return
    
    def start_check(self):
        """
        Opens selected Excel file, starts check from entered starting cell,
        checks text using rules for selected language (English or Simplified Chinese).
        If any errors are found, it prints the error in the cell and changes the cell
        background color to red. Finally, it saves a copy of the file with a new name.
        """
        filename = self.file_entry.get()
        title = 'XXX様映像スタイルガイドツール'
        try:
            wb = openpyxl.load_workbook(filename)
            messagebox.showinfo(title, os.path.basename(filename) + 'を開きます。')
        except OSError:
            messagebox.showinfo(title, '正しいファイルパースを選択してください。')
            return
        ws = wb.active
        max_row = ws.max_row
        start_cell = self.start_cell.get()
        try:
            col = ord(start_cell[0]) - ord("A") + 1
            row = int(start_cell[1])
        except:
            messagebox.showinfo(title, '最初のセルを正しく入力してください。')
            return
        rep = 1
        max_length = 100
        matome_row = 1
        if self.lang.get() == 1:
            title = "Hints: "
            final_cut = "Challenges"
            cut1 = "Hints"
        elif self.lang.get() == 2:
            title = "使用　"
            final_cut = "我们"
            cut1 = "使用"
        for i in range(row, max_row):
            text = ws.cell(row=row,column=col).value
            if rep < 3:
                max_length = 100
            elif rep == 3:
                if self.lang.get() == 1:
                    max_length = 95
                elif self.lang.get() == 2:
                    max_length = 43
            elif rep == 7:
                if self.lang.get() == 1:
                    max_length = 38
                elif self.lang.get() == 2:
                    max_length = 21
            elif rep == 8:
                if self.lang.get() == 1:
                    max_length = 39
                elif self.lang.get() == 2:
                    max_length = 21
            else:
                if self.lang.get() == 1:
                    max_length = 58
                elif self.lang.get() == 2:
                    max_length = 32
            if text:
                text = text.split("\n")
                if final_cut in text:
                    matome_row = row - 1
                for j in text:
                    if len(j) > max_length:
                        ws.cell(row=row,column=col).value += " >" + str(max_length)
                        ws.cell(row=row,column=col).fill = openpyxl.styles.PatternFill("solid", fgColor="FF0000")
                    if "※" in j:
                        ws.cell(row=row,column=col).value += " ※ではなく、＊を使ってください。"
                        ws.cell(row=row,column=col).fill = openpyxl.styles.PatternFill("solid", fgColor="FF0000")
                if rep == 1:
                    if not text[0].startswith(title):
                        ws.cell(row=row,column=col).value += " タイトルのフォーマットは正しくない"
                        ws.cell(row=row,column=col).fill = openpyxl.styles.PatternFill("solid", fgColor="FF0000")
                if rep == 7:
                    if text[0].startswith(cut1):
                        rep -= 1
            row += 1
            rep += 1
        matome_text = ws.cell(row=matome_row, column=col).value
        if matome_text:
            matome_text = matome_text.split("\n")
            if self.lang.get() == 1:
                max_length = 40
            elif self.lang.get() == 2:
                max_length = 21
            for t in matome_text:
                if len(t) > max_length:
                    ws.cell(row=matome_row,column=col).value = ws.cell(row=matome_row,column=col).value[:-3] + " >" + str(max_length)
                    ws.cell(row=matome_row,column=col).fill = openpyxl.styles.PatternFill("solid", fgColor="FF0000")
        newFilename = 'Checked_' + os.path.basename(filename)
        try:
            wb.save(os.path.join(os.path.dirname(filename), newFilename))
            wb.close()
            messagebox.showinfo(title, '終わりました！')      
        except:
            tk.messagebow.showinfo(title, 'エクセルを閉じてからツールを使ってください。')
        return

if __name__ == '__main__':   
    root = tk.Tk()
    app = App(root)
    root.mainloop()  